#!/usr/bin/env python

## PURPOSE: Convert metric_details.tsv to metrics.xlsx.
## USAGE:   python 10a-metrics_excel.py <metrics_file> <uniqreads_file> <genomesize> <ashleysthreshold> <bgthreshold> <wcthreshold> 
## OUTPUT:  metrics.xlsx

import os
import re
import sys
import openpyxl
import pandas as pd
import numpy as np

## VARIABLES
metrics_file = sys.argv[1]
uniqreads_file = sys.argv[2]
genomesize = float(sys.argv[3])
ashleysthreshold = float(sys.argv[4])
bgthreshold = float(sys.argv[5])
wcthreshold = float(sys.argv[6])


## SCRIPT
columns_order = ['Quality', 'Background', 'Duplication_rate', 'Complexity_at_1Gb', 'Median_insert_size', 
                 'Adapters_at_least_50bp', 'Alignment_rate', 'Initial_reads_aligned', 'Processed_reads_aligned', 
                 'Coverage', 'Reads_per_Mb', 'Mean_GC', 'Percent_WC', 'Read_length']
libmetadata = ['Library', 'R', 'C', 'Sample']

metdf = pd.read_table(metrics_file, sep="\t")

metdf_chiponly = metdf.query("not R.isna()").copy()
metdf_sampleonly = metdf_chiponly.query('Sample != ["negativecontrol", "empty"]')[['Sample', 'Quality', 'Background', 'Percent_WC']]

# Good Libraries Per Sample Summary
summary_columns = []
for n in columns_order:
    summary_columns.append("mean_" + n)
    summary_columns.append("sd_" + n)

goodlibs_df = metdf_chiponly.query('Sample != ["negativecontrol", "empty"] & Quality > @ashleysthreshold & Background <= @bgthreshold & Percent_WC <= @wcthreshold')

mean = goodlibs_df.drop(columns=['Library', 'R', 'C']).groupby("Sample").mean()[columns_order].rename(
        columns=(lambda x: "mean_" + x)).reset_index()
std = goodlibs_df.drop(columns=['Library', 'R', 'C']).groupby("Sample").std()[columns_order].rename(
    columns=(lambda x: "sd_" + x)).reset_index()

summary_df = pd.merge(mean, std, on="Sample")[['Sample'] + summary_columns]
summary_df.loc[max(summary_df.index) + 1, "Sample"] = f"Using good-quality libraries only (Quality > {ashleysthreshold}, Background <= {bgthreshold}, Percent_WC <= {wcthreshold})."

# Quality Counts
goodq_counts= goodlibs_df[["Sample", "Quality"]].groupby("Sample").count().rename(columns={'Quality':'Good_libraries'}).reset_index()
poorq_counts = metdf_sampleonly.query('Quality <= @ashleysthreshold | Background > @bgthreshold | Percent_WC > @wcthreshold')[["Sample", "Quality"]].groupby("Sample").count().rename(columns={'Quality':'Poor_libraries'}).reset_index()
empty_counts = metdf_sampleonly.query('Quality.isna()').fillna(1)[["Sample", "Quality"]].groupby("Sample").count().rename(columns={'Quality':'Empty_libraries'}).reset_index()
quality_summary = metdf_sampleonly.drop(columns=['Background', 'Percent_WC']).fillna(1).groupby("Sample").count().rename(columns={'Quality':'All_libraries'}).reset_index()

for n in [goodq_counts, poorq_counts, empty_counts]:
    colname = n.columns[1]
    new = pd.merge(n, quality_summary[["Sample", "All_libraries"]], on="Sample", how='right')
    new[colname + "_Frac"] = round(new[colname] / new["All_libraries"], 3)
    quality_summary = pd.merge(quality_summary, new.drop(columns=["All_libraries"]), on='Sample').fillna(0)

df_uniqreads = pd.read_table(uniqreads_file)
quality_summary = pd.merge(quality_summary, df_uniqreads, on="Sample", how='left')
quality_summary["XCoverage"] = round(quality_summary["Good_UniqReads"] * summary_df['mean_Read_length'].mean() / genomesize, 3)

# Export to Excel
def ws_formating(df, ws):
    for i in list(range(0, len(df.columns))):
        if i < 26:
            letter = chr(i + 65)
        else:
            letter = "A" + chr(i % 26 + 65)

        ws.column_dimensions[letter].width = max(df.iloc[:,i].astype(str).map(len).max(), len(df.columns[i])) + 2
        
        if re.search(("xcoverage"), df.columns[i].lower()):
            for cell in ws[letter]:
                cell.number_format = "0.000"
        elif re.search(("read|^r$|^c$|insert|_libraries$"), df.columns[i].lower()):
            for cell in ws[letter]:
                cell.number_format = "0"
        else:
            for cell in ws[letter]:
                cell.number_format = "0.00%"

with pd.ExcelWriter('metrics.xlsx', engine='openpyxl') as writer:
    summary_df.to_excel(writer, sheet_name="Summary", index=False)  
    metdf[libmetadata + columns_order].to_excel(writer, sheet_name="Details", index=False)
    quality_summary.to_excel(writer, sheet_name="Quality", index=False)
    
wb = openpyxl.load_workbook('metrics.xlsx')
ws_formating(summary_df, wb["Summary"])
wb["Summary"].freeze_panes = "B2"
ws_formating(metdf[libmetadata + columns_order], wb["Details"])
wb["Details"].freeze_panes = "D2"
ws_formating(quality_summary, wb["Quality"])
wb.save('metrics.xlsx')
